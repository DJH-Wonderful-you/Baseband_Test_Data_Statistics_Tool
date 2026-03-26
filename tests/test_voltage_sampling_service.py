from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.core.charge_parser import parse_voltage_sampling_workbook
from src.core.voltage_sampling_service import process_voltage_sampling_statistics


class VoltageSamplingParserTests(unittest.TestCase):
    def test_parse_voltage_sampling_workbook_builds_current_and_preserves_sampling_voltage(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / 'demo.xlsx'
            self._build_demo_workbook(workbook_path)

            dataset = parse_voltage_sampling_workbook(workbook_path, 0.02)

            self.assertEqual(dataset.index_values, [1, 2, 3])
            self.assertEqual(dataset.voltages_v, [3.8, 3.81, 3.82])
            self.assertEqual(dataset.currents_ma, [100.0, 200.0, 50.0])
            self.assertEqual(dataset.extras['分压电压 (V)'], [-0.002, -0.004, -0.001])
            self.assertEqual(dataset.extra_headers_order[0], '分压电压 (V)')
            self.assertEqual(dataset.extras['备注'], ['A', 'B', 'C'])
            self.assertTrue(any('电流方向与预期相反' in warning for warning in dataset.warnings))

    @staticmethod
    def _build_demo_workbook(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['索引', '时间 (s)', '111 电压 (V)', '112 电压 (V)', '备注'])
        start = datetime(2026, 3, 27, 12, 0, 0)
        rows = [
            [1, start, 3.8, -0.002, 'A'],
            [2, start + timedelta(seconds=1), 3.81, -0.004, 'B'],
            [3, start + timedelta(seconds=2), 3.82, -0.001, 'C'],
        ]
        for row in rows:
            sheet.append(row)
        workbook.save(path)


class VoltageSamplingIntegrationTests(unittest.TestCase):
    def test_process_voltage_sampling_statistics_handles_sample_workbook(self) -> None:
        sample_path = Path('test file') / '分压采集测试' / '充电棒 连接星闪.xlsx'
        with tempfile.TemporaryDirectory() as temp_dir:
            result = process_voltage_sampling_statistics([sample_path], Path(temp_dir), 20)

            self.assertEqual(result.total, 1)
            self.assertEqual(result.success, 1)
            self.assertEqual(result.failed, 0)
            output_path = result.items[0].output_path
            self.assertIsNotNone(output_path)

            source_workbook = load_workbook(sample_path, data_only=True)
            source_sheet = source_workbook.active
            output_workbook = load_workbook(output_path, data_only=True)
            output_sheet = output_workbook.active

            headers = [output_sheet.cell(row=1, column=index).value for index in range(1, 7)]
            self.assertEqual(
                headers,
                ['索引', '日期', '时间 (s)', '电流 (mA)', '电压（V）', '分压电压 (V)'],
            )
            self.assertEqual(output_sheet.cell(row=1, column=8).value, '充电曲线测试')
            self.assertEqual(len(getattr(output_sheet, '_charts', [])), 1)

            source_sampling_voltage = source_sheet.cell(row=2, column=4).value
            self.assertAlmostEqual(output_sheet.cell(row=2, column=6).value, source_sampling_voltage)
            expected_current_ma = source_sampling_voltage / 0.02 * 1000.0
            self.assertAlmostEqual(output_sheet.cell(row=2, column=4).value, expected_current_ma)


if __name__ == '__main__':
    unittest.main()
