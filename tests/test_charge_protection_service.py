from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from src.core.charge_protection_service import (
    HIGH_LOW_MODE,
    HIGH_ONLY_MODE,
    LOW_ONLY_MODE,
    build_charge_protection_dataset,
    compute_charge_protection_metrics,
    process_charge_protection_statistics,
)
from src.core.errors import AppError
from src.core.models import ChargeDataset, ChargeProtectionDataset


def _build_charge_protection_dataset(
    *,
    stem: str,
    currents_ma: list[float | None],
    cell_temps_c: list[float | None],
    env_temps_c: list[float | None] | None = None,
) -> ChargeProtectionDataset:
    row_count = len(currents_ma)
    if env_temps_c is None:
        env_temps_c = [25.0] * row_count
    datetimes = [datetime(2026, 3, 26, 12, 0, 0) + timedelta(seconds=index) for index in range(row_count)]
    return ChargeProtectionDataset(
        source_path=Path(f"{stem}.xlsx"),
        stem=stem,
        index_values=list(range(1, row_count + 1)),
        datetimes=datetimes,
        date_strings=[value.strftime("%Y-%m-%d") for value in datetimes],
        time_strings=[value.strftime("%H:%M:%S") for value in datetimes],
        currents_ma=currents_ma,
        voltages_v=[3.7] * row_count,
        env_temps_c=env_temps_c,
        cell_temps_c=cell_temps_c,
        warnings=[],
    )


class ChargeProtectionServiceTests(unittest.TestCase):
    def test_build_charge_protection_dataset_promotes_cell_temp_from_extras(self) -> None:
        row_count = 4
        datetimes = [datetime(2026, 3, 26, 12, 0, 0) + timedelta(seconds=index) for index in range(row_count)]
        dataset = ChargeDataset(
            source_path=Path("demo.xlsx"),
            stem="demo",
            index_values=[1, 2, 3, 4],
            datetimes=datetimes,
            date_strings=[value.strftime("%Y-%m-%d") for value in datetimes],
            time_strings=[value.strftime("%H:%M:%S") for value in datetimes],
            currents_ma=[10.0, 9.0, 0.0, 5.0],
            voltages_v=[3.8, 3.8, 3.8, 3.8],
            pen_temps_c=[None, None, None, None],
            env_temps_c=[20.0, 20.5, 21.0, 21.5],
            extras={"104 电芯温度 温度 (°C)": [48.0, 49.0, 50.0, 45.0]},
            extra_headers_order=["104 电芯温度 温度 (°C)"],
            has_temperature_data=False,
            warnings=[],
        )

        protection_dataset = build_charge_protection_dataset(dataset)

        self.assertEqual(protection_dataset.env_temps_c, [20.0, 20.5, 21.0, 21.5])
        self.assertEqual(protection_dataset.cell_temps_c, [48.0, 49.0, 50.0, 45.0])

    def test_compute_charge_protection_metrics_selects_longest_interval_and_last_resume(self) -> None:
        dataset = _build_charge_protection_dataset(
            stem="demo 高低温充电保护测试",
            currents_ma=[5, 5, 0, 0, 0, 5, 5, 5, 0, 0, 5, 5, 0, 0, 5],
            cell_temps_c=[30, 50, 51, 52, 53, 35, 36, 52, 49, 48, 10, -10, -8, -5, 5],
        )

        metrics = compute_charge_protection_metrics(dataset)

        self.assertEqual(metrics.mode, HIGH_LOW_MODE)
        self.assertEqual(metrics.high_protect_temp_c, 50)
        self.assertEqual(metrics.high_resume_temp_c, 48)
        self.assertEqual(metrics.low_protect_temp_c, -10)
        self.assertEqual(metrics.low_resume_temp_c, -5)

    def test_compute_charge_protection_metrics_open_interval_sets_warning_and_missing_resume(self) -> None:
        dataset = _build_charge_protection_dataset(
            stem="demo 高温充电保护测试",
            currents_ma=[5, 5, 0, 0, 0],
            cell_temps_c=[30, 50, 51, 52, 53],
        )

        metrics = compute_charge_protection_metrics(dataset)

        self.assertEqual(metrics.mode, HIGH_ONLY_MODE)
        self.assertEqual(metrics.high_protect_temp_c, 50)
        self.assertIsNone(metrics.high_resume_temp_c)
        self.assertTrue(any("未检测到终点" in warning for warning in dataset.warnings))

    def test_compute_charge_protection_metrics_ignores_false_pause(self) -> None:
        dataset = _build_charge_protection_dataset(
            stem="demo 低温充电保护测试",
            currents_ma=[5, 5, 0, 0, 5, 5, 0, 0, 5],
            cell_temps_c=[25, 25, 24, 24, 10, -8, -7, -3, 8],
        )

        metrics = compute_charge_protection_metrics(dataset)

        self.assertEqual(metrics.mode, LOW_ONLY_MODE)
        self.assertEqual(metrics.low_protect_temp_c, -8)
        self.assertEqual(metrics.low_resume_temp_c, -3)
        self.assertTrue(any("虚假停充区间" in warning for warning in dataset.warnings))

    def test_compute_charge_protection_metrics_raises_when_high_low_filename_mismatch(self) -> None:
        dataset = _build_charge_protection_dataset(
            stem="demo 高低温充电保护测试",
            currents_ma=[5, 5, 0, 0, 5],
            cell_temps_c=[25, 50, 51, 49, 30],
        )

        with self.assertRaises(AppError) as context:
            compute_charge_protection_metrics(dataset)
        self.assertIn("未同时检测到高温区间和低温区间", str(context.exception))

    def test_compute_charge_protection_metrics_raises_when_interval_matches_both_high_and_low(self) -> None:
        dataset = _build_charge_protection_dataset(
            stem="demo 充电保护测试",
            currents_ma=[5, 5, 0, 0, 5],
            cell_temps_c=[25, 50, -2, 49, 30],
        )

        with self.assertRaises(AppError) as context:
            compute_charge_protection_metrics(dataset)
        self.assertIn("同时满足高温和低温判定条件", str(context.exception))

    def test_compute_charge_protection_metrics_raises_when_no_start(self) -> None:
        dataset = _build_charge_protection_dataset(
            stem="demo 充电保护测试",
            currents_ma=[5, 5, 5, 5],
            cell_temps_c=[25, 26, 27, 28],
        )

        with self.assertRaises(AppError) as context:
            compute_charge_protection_metrics(dataset)
        self.assertIn("未检测到任何停充区间起点", str(context.exception))


class ChargeProtectionIntegrationTests(unittest.TestCase):
    def test_process_charge_protection_statistics_handles_high_temperature_sample(self) -> None:
        sample_path = Path("test file") / "充电保护测试" / "FL01 dvt-a高温充电保护测试0323.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            result = process_charge_protection_statistics([sample_path], Path(temp_dir))

            self.assertEqual(result.total, 1)
            self.assertEqual(result.success, 1)
            self.assertEqual(result.failed, 0)
            output_path = result.items[0].output_path
            self.assertIsNotNone(output_path)
            workbook = load_workbook(output_path)
            sheet = workbook.active
            headers = [sheet.cell(row=1, column=index).value for index in range(1, 8)]
            self.assertEqual(
                headers,
                ["索引", "日期", "时间 (s)", "电流 (mA)", "电压 (V)", "环境温度 (°C)", "电芯温度 (°C)"],
            )
            self.assertEqual(sheet.cell(row=1, column=9).value, "高温充电保护测试")
            self.assertEqual(len(getattr(sheet, "_charts", [])), 1)

    def test_process_charge_protection_statistics_handles_low_temperature_sample(self) -> None:
        sample_path = Path("test file") / "充电保护测试" / "FL01 dvt-a低温充电保护测试0323.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            result = process_charge_protection_statistics([sample_path], Path(temp_dir))

            self.assertEqual(result.total, 1)
            self.assertEqual(result.success, 1)
            self.assertEqual(result.failed, 0)
            output_path = result.items[0].output_path
            self.assertIsNotNone(output_path)
            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=1, column=9).value, "低温充电保护测试")
            self.assertEqual(len(getattr(sheet, "_charts", [])), 1)


if __name__ == "__main__":
    unittest.main()
