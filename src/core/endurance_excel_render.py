from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ChargeDataset


DEFAULT_CHART_WIDTH = 16
DEFAULT_CHART_HEIGHT = 10
PLOT_AREA_X = 0.0
PLOT_AREA_Y = 0.0
PLOT_AREA_W = 0.9
PLOT_AREA_H = 0.7
CHART_TITLE_SIZE = 1600


@dataclass(slots=True)
class _ColumnDef:
    header: str
    values: list[Any]
    number_format: str | None = None


@dataclass(slots=True)
class IndicatorLogRow:
    log_datetime: datetime
    level: int
    mapped_voltage_v: float | None
    log_date: str | None = None


def _format_duration(value: timedelta) -> str:
    total_seconds = int(value.total_seconds())
    if total_seconds < 0:
        total_seconds = abs(total_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def _apply_plot_area_layout(chart: LineChart) -> None:
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="outer",
            x=PLOT_AREA_X,
            y=PLOT_AREA_Y,
            w=PLOT_AREA_W,
            h=PLOT_AREA_H,
            xMode="factor",
            yMode="factor",
            wMode="factor",
            hMode="factor",
        )
    )


def _base_columns(dataset: ChargeDataset) -> list[_ColumnDef]:
    columns: list[_ColumnDef] = []
    if any(value is not None for value in dataset.index_values):
        columns.append(_ColumnDef("索引", dataset.index_values))
    columns.append(_ColumnDef("日期", dataset.date_strings))
    columns.append(_ColumnDef("时间 (s)", dataset.datetimes, number_format="hh:mm:ss"))
    columns.append(_ColumnDef("电流 (mA)", dataset.currents_ma, number_format="0.000"))
    columns.append(_ColumnDef("电压（V）", dataset.voltages_v, number_format="0.000"))
    if dataset.has_temperature_data:
        columns.append(_ColumnDef("笔壳温度 (°C)", dataset.pen_temps_c, number_format="0.000"))
        columns.append(_ColumnDef("环境温度 (°C)", dataset.env_temps_c, number_format="0.000"))
    for header in dataset.extra_headers_order:
        columns.append(_ColumnDef(header, dataset.extras.get(header, [])))
    return columns


def _write_base_headers(ws: Any, columns: list[_ColumnDef]) -> None:
    header_font = Font(bold=True)
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column.header)
        cell.font = header_font
        if column.header in {"日期", "时间 (s)"}:
            width = 14
        elif "温度" in column.header:
            width = 16
        else:
            width = 14
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_column_values(ws: Any, columns: list[_ColumnDef], total_rows: int) -> None:
    for row_offset in range(total_rows):
        row_idx = row_offset + 2
        for col_idx, column in enumerate(columns, start=1):
            value = column.values[row_offset] if row_offset < len(column.values) else None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if column.number_format and value is not None:
                cell.number_format = column.number_format


def _write_duration_table(ws: Any, start_col: int, duration: timedelta) -> int:
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")

    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
    title = ws.cell(row=1, column=start_col, value="续航时长统计")
    title.font = title_font
    title.alignment = center

    header_left = ws.cell(row=2, column=start_col, value="测试项")
    header_right = ws.cell(row=2, column=start_col + 1, value="数据")
    for header_cell in (header_left, header_right):
        header_cell.font = title_font
        header_cell.alignment = center
        header_cell.fill = header_fill

    ws.cell(row=3, column=start_col, value="续航时长")
    value_cell = ws.cell(row=3, column=start_col + 1, value=_format_duration(duration))
    value_cell.alignment = center
    ws.column_dimensions[get_column_letter(start_col)].width = 16
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 28
    return 3


def _add_current_voltage_chart(
    ws: Any,
    *,
    title: str,
    first_data_row: int,
    last_data_row: int,
    time_col: int,
    current_col: int,
    voltage_col: int,
    anchor_cell: str,
) -> None:
    primary_x_axis_id = 500
    primary_y_axis_id = 100
    secondary_y_axis_id = 200

    chart_left = LineChart()
    chart_left.title = title
    chart_left.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=CHART_TITLE_SIZE, b=True)
    chart_left.x_axis.axId = primary_x_axis_id
    chart_left.y_axis.axId = primary_y_axis_id
    chart_left.y_axis.title = "电流"
    chart_left.y_axis.axPos = "l"
    chart_left.y_axis.delete = False
    chart_left.y_axis.tickLblPos = "nextTo"
    chart_left.y_axis.number_format = "0.000"
    chart_left.y_axis.crosses = "min"
    chart_left.y_axis.crossAx = primary_x_axis_id
    chart_left.x_axis.title = "时间"
    chart_left.x_axis.axPos = "b"
    chart_left.x_axis.delete = False
    chart_left.x_axis.crossAx = primary_y_axis_id
    chart_left.x_axis.number_format = "hh:mm:ss"
    chart_left.x_axis.tickLblPos = "low"
    chart_left.legend.position = "b"
    chart_left.width = DEFAULT_CHART_WIDTH
    chart_left.height = DEFAULT_CHART_HEIGHT
    _apply_plot_area_layout(chart_left)

    current_data = Reference(ws, min_col=current_col, min_row=1, max_col=current_col, max_row=last_data_row)
    categories = Reference(ws, min_col=time_col, min_row=first_data_row, max_col=time_col, max_row=last_data_row)
    chart_left.add_data(current_data, titles_from_data=True)
    chart_left.set_categories(categories)

    chart_right = LineChart()
    chart_right.x_axis.axId = primary_x_axis_id
    chart_right.y_axis.axId = secondary_y_axis_id
    chart_right.y_axis.title = "电压"
    chart_right.y_axis.axPos = "r"
    chart_right.y_axis.delete = False
    chart_right.y_axis.tickLblPos = "nextTo"
    chart_right.y_axis.number_format = "0.000"
    chart_right.y_axis.crosses = "max"
    chart_right.y_axis.crossAx = primary_x_axis_id
    chart_right.y_axis.majorGridlines = None
    voltage_data = Reference(ws, min_col=voltage_col, min_row=1, max_col=voltage_col, max_row=last_data_row)
    chart_right.add_data(voltage_data, titles_from_data=True)
    chart_left += chart_right
    ws.add_chart(chart_left, anchor_cell)


def _add_voltage_level_chart(
    ws: Any,
    *,
    title: str,
    first_data_row: int,
    last_data_row: int,
    time_col: int,
    voltage_col: int,
    level_col: int,
    anchor_cell: str,
) -> None:
    primary_x_axis_id = 700
    primary_y_axis_id = 300
    secondary_y_axis_id = 400

    chart_left = LineChart()
    chart_left.title = title
    chart_left.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=CHART_TITLE_SIZE, b=True)
    chart_left.x_axis.axId = primary_x_axis_id
    chart_left.y_axis.axId = primary_y_axis_id
    chart_left.y_axis.title = "电压"
    chart_left.y_axis.axPos = "l"
    chart_left.y_axis.delete = False
    chart_left.y_axis.tickLblPos = "nextTo"
    chart_left.y_axis.number_format = "0.000"
    chart_left.y_axis.crosses = "min"
    chart_left.y_axis.crossAx = primary_x_axis_id
    chart_left.x_axis.title = "时间"
    chart_left.x_axis.axPos = "b"
    chart_left.x_axis.delete = False
    chart_left.x_axis.crossAx = primary_y_axis_id
    chart_left.x_axis.number_format = "hh:mm:ss"
    chart_left.x_axis.tickLblPos = "low"
    chart_left.legend.position = "b"
    chart_left.width = DEFAULT_CHART_WIDTH
    chart_left.height = DEFAULT_CHART_HEIGHT
    _apply_plot_area_layout(chart_left)

    voltage_data = Reference(ws, min_col=voltage_col, min_row=1, max_col=voltage_col, max_row=last_data_row)
    categories = Reference(ws, min_col=time_col, min_row=first_data_row, max_col=time_col, max_row=last_data_row)
    chart_left.add_data(voltage_data, titles_from_data=True)
    chart_left.set_categories(categories)

    chart_right = LineChart()
    chart_right.x_axis.axId = primary_x_axis_id
    chart_right.y_axis.axId = secondary_y_axis_id
    chart_right.y_axis.title = "电量"
    chart_right.y_axis.axPos = "r"
    chart_right.y_axis.delete = False
    chart_right.y_axis.tickLblPos = "nextTo"
    chart_right.y_axis.number_format = "0"
    chart_right.y_axis.crosses = "max"
    chart_right.y_axis.crossAx = primary_x_axis_id
    chart_right.y_axis.majorGridlines = None
    level_data = Reference(ws, min_col=level_col, min_row=1, max_col=level_col, max_row=last_data_row)
    chart_right.add_data(level_data, titles_from_data=True)
    chart_left += chart_right
    ws.add_chart(chart_left, anchor_cell)


def _add_level_only_chart(
    ws: Any,
    *,
    title: str,
    first_data_row: int,
    last_data_row: int,
    time_col: int,
    level_col: int,
    anchor_cell: str,
) -> None:
    primary_x_axis_id = 900
    primary_y_axis_id = 500

    chart = LineChart()
    chart.title = title
    chart.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=CHART_TITLE_SIZE, b=True)
    chart.x_axis.axId = primary_x_axis_id
    chart.y_axis.axId = primary_y_axis_id
    chart.y_axis.title = "电量"
    chart.y_axis.axPos = "l"
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.number_format = "0"
    chart.y_axis.crosses = "min"
    chart.y_axis.crossAx = primary_x_axis_id
    chart.x_axis.title = "时间"
    chart.x_axis.axPos = "b"
    chart.x_axis.delete = False
    chart.x_axis.crossAx = primary_y_axis_id
    chart.x_axis.number_format = "hh:mm:ss"
    chart.x_axis.tickLblPos = "low"
    chart.legend.position = "b"
    chart.width = DEFAULT_CHART_WIDTH
    chart.height = DEFAULT_CHART_HEIGHT
    _apply_plot_area_layout(chart)

    level_data = Reference(ws, min_col=level_col, min_row=1, max_col=level_col, max_row=last_data_row)
    categories = Reference(ws, min_col=time_col, min_row=first_data_row, max_col=time_col, max_row=last_data_row)
    chart.add_data(level_data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor_cell)


def _center_cells_with_data(ws: Any) -> None:
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and value.strip() == "":
                continue
            cell.alignment = center


def _save_workbook(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_dir = output_path.parent / ".temp"
    temp_dir.mkdir(exist_ok=True)
    original_tempdir = tempfile.tempdir
    try:
        tempfile.tempdir = str(temp_dir)
        workbook.save(output_path)
    finally:
        tempfile.tempdir = original_tempdir
        try:
            for temp_file in temp_dir.glob("openpyxl.*"):
                temp_file.unlink(missing_ok=True)
            if not any(temp_dir.iterdir()):
                temp_dir.rmdir()
        except Exception:
            pass


def render_endurance_duration_workbook(
    dataset: ChargeDataset,
    endurance_duration: timedelta,
    output_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统计结果"

    columns = _base_columns(dataset)
    _write_base_headers(sheet, columns)
    _write_column_values(sheet, columns, dataset.row_count())

    summary_col = len(columns) + 2
    table_end_row = _write_duration_table(sheet, summary_col, endurance_duration)

    header_to_col = {column.header: idx for idx, column in enumerate(columns, start=1)}
    time_col = header_to_col["时间 (s)"]
    current_col = header_to_col["电流 (mA)"]
    voltage_col = header_to_col["电压（V）"]
    chart_anchor = f"{get_column_letter(summary_col)}{table_end_row + 2}"
    _add_current_voltage_chart(
        sheet,
        title=dataset.stem,
        first_data_row=2,
        last_data_row=dataset.row_count() + 1,
        time_col=time_col,
        current_col=current_col,
        voltage_col=voltage_col,
        anchor_cell=chart_anchor,
    )

    _center_cells_with_data(sheet)
    _save_workbook(workbook, output_path)


def render_endurance_indicator_workbook(
    dataset: ChargeDataset,
    log_rows: list[IndicatorLogRow],
    include_log_date: bool,
    endurance_duration: timedelta,
    output_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统计结果"

    base_columns = _base_columns(dataset)
    _write_base_headers(sheet, base_columns)
    total_rows = max(dataset.row_count(), len(log_rows))
    _write_column_values(sheet, base_columns, total_rows)

    log_start_col = len(base_columns) + 1
    log_columns = ["log-时间 (s)", "电量 (%)", "映射电压 (V)"]
    if include_log_date:
        log_columns.insert(0, "log-日期")
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    for offset, header in enumerate(log_columns):
        col_idx = log_start_col + offset
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        width = 14 if "时间" in header or "日期" in header else 12
        if "映射电压" in header:
            width = 14
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset, log_row in enumerate(log_rows):
        row_idx = row_offset + 2
        value_offset = 0
        if include_log_date:
            date_cell = sheet.cell(row=row_idx, column=log_start_col, value=log_row.log_date)
            date_cell.number_format = "@"
            value_offset = 1

        time_col_idx = log_start_col + value_offset
        time_cell = sheet.cell(row=row_idx, column=time_col_idx, value=log_row.log_datetime)
        time_cell.number_format = "hh:mm:ss"

        level_col_idx = time_col_idx + 1
        sheet.cell(row=row_idx, column=level_col_idx, value=log_row.level)

        voltage_col_idx = level_col_idx + 1
        voltage_cell = sheet.cell(row=row_idx, column=voltage_col_idx, value=log_row.mapped_voltage_v)
        if log_row.mapped_voltage_v is not None:
            voltage_cell.number_format = "0.000"

    summary_col = log_start_col + len(log_columns) + 2
    table_end_row = _write_duration_table(sheet, summary_col, endurance_duration)

    if log_rows:
        time_col_idx = log_start_col + (1 if include_log_date else 0)
        level_col_idx = time_col_idx + 1
        voltage_col_idx = level_col_idx + 1
        chart_anchor = f"{get_column_letter(summary_col)}{table_end_row + 2}"
        _add_voltage_level_chart(
            sheet,
            title=dataset.stem,
            first_data_row=2,
            last_data_row=len(log_rows) + 1,
            time_col=time_col_idx,
            voltage_col=voltage_col_idx,
            level_col=level_col_idx,
            anchor_cell=chart_anchor,
        )

    _center_cells_with_data(sheet)
    _save_workbook(workbook, output_path)


def render_endurance_single_log_workbook(
    *,
    file_stem: str,
    log_rows: list[IndicatorLogRow],
    include_voltage: bool,
    endurance_duration: timedelta,
    output_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统计结果"

    log_columns = ["log-时间 (s)", "电量 (%)"]
    if include_voltage:
        log_columns.append("映射电压 (V)")

    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    for offset, header in enumerate(log_columns):
        col_idx = offset + 1
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        width = 14 if "时间" in header or "日期" in header else 12
        if "电压" in header:
            width = 14
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset, log_row in enumerate(log_rows):
        row_idx = row_offset + 2
        time_cell = sheet.cell(row=row_idx, column=1, value=log_row.log_datetime)
        time_cell.number_format = "hh:mm:ss"

        sheet.cell(row=row_idx, column=2, value=log_row.level)
        if include_voltage:
            voltage_cell = sheet.cell(row=row_idx, column=3, value=log_row.mapped_voltage_v)
            if log_row.mapped_voltage_v is not None:
                voltage_cell.number_format = "0.000"

    summary_col = len(log_columns) + 2
    table_end_row = _write_duration_table(sheet, summary_col, endurance_duration)

    if log_rows:
        chart_anchor = f"{get_column_letter(summary_col)}{table_end_row + 2}"
        _add_level_only_chart(
            sheet,
            title=file_stem,
            first_data_row=2,
            last_data_row=len(log_rows) + 1,
            time_col=1,
            level_col=2,
            anchor_cell=chart_anchor,
        )

    _center_cells_with_data(sheet)
    _save_workbook(workbook, output_path)
