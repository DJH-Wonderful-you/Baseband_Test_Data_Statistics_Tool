from __future__ import annotations

import csv
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .errors import AppError
from .models import ChargeDataset


class _CellProxy:
    def __init__(self, value: Any) -> None:
        self.value = value


class _SheetProxy:
    def __init__(self, values: list[list[Any]]) -> None:
        self._values = values
        self.max_row = len(values)
        self.max_column = max((len(row) for row in values), default=0)

    def cell(self, row: int, column: int) -> _CellProxy:
        if row <= 0 or column <= 0 or row > self.max_row:
            return _CellProxy(None)
        row_values = self._values[row - 1]
        if column > len(row_values):
            return _CellProxy(None)
        return _CellProxy(row_values[column - 1])


def _load_sheet(path: Path) -> tuple[Any, datetime]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        # Some files are actually xlsx content but saved/renamed as .xls.
        # Detect by zip signature and parse them as openpyxl workbooks.
        if zipfile.is_zipfile(path):
            with path.open("rb") as file_obj:
                workbook = load_workbook(file_obj, data_only=True)
            return workbook[workbook.sheetnames[0]], workbook.epoch
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise AppError(
                "XLS_DEPENDENCY_MISSING",
                "检测到 .xls 文件，但缺少解析依赖 xlrd，请先安装 requirements.txt 后重试",
                detail=str(path),
            ) from exc
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        rows: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            row_values: list[Any] = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode).replace(microsecond=0)
                row_values.append(value)
            rows.append(row_values)
        epoch = datetime(1899, 12, 30) if workbook.datemode == 0 else datetime(1904, 1, 1)
        return _SheetProxy(rows), epoch

    workbook = load_workbook(path, data_only=True)
    return workbook[workbook.sheetnames[0]], workbook.epoch


def _normalize_header(text: str) -> str:
    normalized = text.strip().lower()
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized = normalized.replace(" ", "")
    return normalized


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    if _is_empty(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _safe_str(value).replace(",", "")
    match = re.search(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def _parse_date_only(value: Any, epoch: datetime) -> date | None:
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch=epoch)
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
        return None
    text = _safe_str(value)
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if not match:
        return None
    year, month, day = map(int, match.groups())
    return date(year, month, day)


def _parse_datetime_with_truncation(
    value: Any, epoch: datetime, fallback_date: date | None = None
) -> datetime | None:
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch=epoch)
        if isinstance(converted, datetime):
            return converted.replace(microsecond=0)
        if isinstance(converted, date):
            return datetime.combine(converted, time.min)
        if isinstance(converted, time):
            if fallback_date is None:
                return None
            return datetime.combine(fallback_date, converted).replace(microsecond=0)
        return None

    text = _safe_str(value)
    full_match = re.search(
        r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})[:：](\d{1,2})[:：](\d{1,2})",
        text,
    )
    if full_match:
        year, month, day, hour, minute, second = map(int, full_match.groups())
        return datetime(year, month, day, hour, minute, second)

    time_match = re.search(r"(\d{1,2})[:：](\d{1,2})[:：](\d{1,2})", text)
    if time_match and fallback_date is not None:
        hour, minute, second = map(int, time_match.groups())
        return datetime.combine(fallback_date, time(hour, minute, second))
    return None


def _find_header_row(ws: Any) -> int:
    keyword_set = ("时间", "电流", "电压", "索引", "笔壳", "环境")
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        score = 0
        non_empty = 0
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if _is_empty(value):
                continue
            non_empty += 1
            text = _safe_str(value)
            if any(keyword in text for keyword in keyword_set):
                score += 1
        if non_empty == 0:
            continue
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def _read_headers(ws: Any, header_row: int) -> dict[int, str]:
    headers: dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if _is_empty(value):
            continue
        headers[col_idx] = _safe_str(value)
    return headers


def _find_column(headers: dict[int, str], predicate: Any) -> int | None:
    for col_idx, text in headers.items():
        if predicate(text):
            return col_idx
    return None


def _build_extra_headers(headers: dict[int, str], excluded_columns: set[int]) -> list[tuple[int, str]]:
    extras: list[tuple[int, str]] = []
    existing: set[str] = set()
    for col_idx, text in headers.items():
        if col_idx in excluded_columns:
            continue
        candidate = text
        if candidate in existing:
            candidate = f"{candidate}_{col_idx}"
        existing.add(candidate)
        extras.append((col_idx, candidate))
    return extras


def _is_next_second(current_second: int, target_second: int, step: int) -> bool:
    return (current_second + step) % 60 == target_second


def _normalize_duplicate_second_records(
    records: list[dict[str, Any]],
    source_label: str,
    path: Path,
) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if len(records) < 2:
        return records, warnings

    fixed_counter = 0
    for idx in range(len(records) - 1):
        first = records[idx]
        second = records[idx + 1]
        if first["dt"] != second["dt"]:
            continue
        if idx == 0 or idx + 2 >= len(records):
            continue

        prev_dt = records[idx - 1]["dt"]
        duplicate_dt = first["dt"]
        next_dt = records[idx + 2]["dt"]
        missing_before = (
            _is_next_second(prev_dt.second, duplicate_dt.second, 2)
            and _is_next_second(duplicate_dt.second, next_dt.second, 1)
        )
        missing_after = (
            _is_next_second(prev_dt.second, duplicate_dt.second, 1)
            and _is_next_second(duplicate_dt.second, next_dt.second, 2)
        )
        if missing_before:
            fixed_dt = duplicate_dt - timedelta(seconds=1)
            first["dt"] = fixed_dt
            fixed_counter += 1
            continue
        if missing_after:
            fixed_dt = duplicate_dt + timedelta(seconds=1)
            second["dt"] = fixed_dt
            fixed_counter += 1

    deduplicated: list[dict[str, Any]] = []
    seen_datetimes: set[datetime] = set()
    duplicate_counter = 0
    for record in records:
        dt_value = record["dt"]
        if dt_value in seen_datetimes:
            duplicate_counter += 1
            continue
        seen_datetimes.add(dt_value)
        deduplicated.append(record)

    if fixed_counter > 0:
        warnings.append(f"{source_label} 文件 {path.name} 共修正 {fixed_counter} 个重复秒级时间点")
    if duplicate_counter > 0:
        warnings.append(f"{source_label} 存在 {duplicate_counter} 个重复秒级时间点，已保留首次出现的数据")
    return deduplicated, warnings


def parse_charge_workbook(
    path: Path,
    require_voltage: bool,
    normalize_duplicate_seconds: bool = False,
    require_current: bool = True,
) -> ChargeDataset:
    sheet, epoch = _load_sheet(path)
    header_row = _find_header_row(sheet)
    headers = _read_headers(sheet, header_row)
    if not headers:
        raise AppError("EMPTY_HEADER", "未检测到有效表头", detail=str(path))

    time_col = _find_column(
        headers,
        lambda h: "时间" in h and "log" not in _normalize_header(h),
    )
    current_col = _find_column(
        headers,
        lambda h: "电流" in h and "log" not in _normalize_header(h),
    )
    voltage_col = _find_column(
        headers,
        lambda h: ("电压" in h or "vbat" in _normalize_header(h))
        and "映射" not in h
        and "log" not in _normalize_header(h),
    )
    index_col = _find_column(headers, lambda h: "索引" in h)
    date_col = _find_column(headers, lambda h: "日期" in h and "log" not in _normalize_header(h))
    pen_col = _find_column(headers, lambda h: "笔壳" in h and "温度" in h)
    env_col = _find_column(headers, lambda h: "环境" in h and "温度" in h)
    has_temperature_data = pen_col is not None and env_col is not None

    if time_col is None:
        raise AppError("MISSING_TIME", "未检测到“时间 (s)”相关列", detail=str(path))
    if require_current and current_col is None:
        raise AppError("MISSING_CURRENT", "未检测到“电流”相关列", detail=str(path))
    if require_voltage and voltage_col is None:
        raise AppError("MISSING_VOLTAGE", "未检测到“电压”相关列", detail=str(path))

    excluded_columns = {time_col, current_col}
    if index_col is not None:
        excluded_columns.add(index_col)
    if date_col is not None:
        excluded_columns.add(date_col)
    if voltage_col is not None:
        excluded_columns.add(voltage_col)
    if has_temperature_data and pen_col is not None:
        excluded_columns.add(pen_col)
    if has_temperature_data and env_col is not None:
        excluded_columns.add(env_col)
    extra_headers = _build_extra_headers(headers, excluded_columns)

    row_records: list[dict[str, Any]] = []
    warnings: list[str] = []
    point_index = 0

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        time_value = sheet.cell(row=row_idx, column=time_col).value
        if _is_empty(time_value):
            continue
        point_index += 1
        fallback_date = None
        if date_col is not None:
            date_value = sheet.cell(row=row_idx, column=date_col).value
            fallback_date = _parse_date_only(date_value, epoch)
        dt_value = _parse_datetime_with_truncation(time_value, epoch, fallback_date)
        if dt_value is None:
            raise AppError(
                "INVALID_DATETIME",
                "时间列存在无法解析的数据",
                detail=f"{path.name} row={row_idx} value={time_value}",
            )

        raw_index_value = sheet.cell(row=row_idx, column=index_col).value if index_col else None
        parsed_index = int(raw_index_value) if isinstance(raw_index_value, int) else _to_float(raw_index_value)
        row_records.append(
            {
                "point_index": point_index,
                "index_value": int(parsed_index) if parsed_index is not None else None,
                "dt": dt_value,
                "current_raw": sheet.cell(row=row_idx, column=current_col).value if current_col is not None else None,
                "voltage_raw": sheet.cell(row=row_idx, column=voltage_col).value if voltage_col is not None else None,
                "pen_temp_c": _to_float(sheet.cell(row=row_idx, column=pen_col).value) if pen_col else None,
                "env_temp_c": _to_float(sheet.cell(row=row_idx, column=env_col).value) if env_col else None,
                "extras": {
                    extra_name: sheet.cell(row=row_idx, column=extra_col).value
                    for extra_col, extra_name in extra_headers
                },
            }
        )

    if not row_records:
        raise AppError("NO_DATA", "未检测到有效数据行", detail=str(path))

    if normalize_duplicate_seconds:
        row_records, duplicate_warnings = _normalize_duplicate_second_records(row_records, "Excel", path)
        warnings.extend(duplicate_warnings)

    index_values: list[int | None] = []
    datetimes: list[datetime] = []
    date_strings: list[str] = []
    time_strings: list[str] = []
    current_raw_values: list[Any] = []
    voltage_raw_values: list[Any] = []
    pen_temps_c: list[float | None] = []
    env_temps_c: list[float | None] = []
    extras: dict[str, list[Any]] = {header: [] for _, header in extra_headers}
    for record in row_records:
        dt_value = record["dt"]
        index_values.append(record["index_value"])
        datetimes.append(dt_value)
        date_strings.append(dt_value.strftime("%Y-%m-%d"))
        time_strings.append(dt_value.strftime("%H:%M:%S"))
        current_raw_values.append(record["current_raw"])
        voltage_raw_values.append(record["voltage_raw"])
        pen_temps_c.append(record["pen_temp_c"])
        env_temps_c.append(record["env_temp_c"])
        for _, extra_name in extra_headers:
            extras[extra_name].append(record["extras"].get(extra_name))

    currents_raw = [_to_float(value) for value in current_raw_values]
    voltages_v = [_to_float(value) for value in voltage_raw_values]

    if current_col is not None:
        current_header_text = headers[current_col]
        current_header_norm = _normalize_header(current_header_text)
        current_factor = 1000.0
        if "ma" in current_header_norm:
            current_factor = 1.0
        elif "(a)" in current_header_norm or "电流(a)" in current_header_norm:
            current_factor = 1000.0
        currents_ma = [
            (value * current_factor if value is not None else None) for value in currents_raw
        ]
        valid_current_values = [value for value in currents_ma if value is not None]
        if valid_current_values and max(valid_current_values, key=lambda value: abs(value)) < 0:
            currents_ma = [(-value if value is not None else None) for value in currents_ma]
            warnings.append("检测到电流方向与预期相反，已将整列电流取相反数后再进行后续统计")
    else:
        currents_ma = [None for _ in currents_raw]

    return ChargeDataset(
        source_path=path,
        stem=path.stem,
        index_values=index_values,
        datetimes=datetimes,
        date_strings=date_strings,
        time_strings=time_strings,
        currents_ma=currents_ma,
        voltages_v=voltages_v,
        pen_temps_c=pen_temps_c,
        env_temps_c=env_temps_c,
        extras=extras,
        extra_headers_order=[header for _, header in extra_headers],
        has_temperature_data=has_temperature_data,
        warnings=warnings,
    )


def parse_time_value_csv(path: Path) -> tuple[dict[datetime, Any], list[str], str | None]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if not reader.fieldnames:
            raise AppError("CSV_EMPTY", "CSV 文件为空或缺少表头", detail=str(path))

        field_map: dict[str, str] = {}
        for field in reader.fieldnames:
            normalized = _normalize_header(field)
            field_map[field] = normalized

        datetime_field = next(
            (
                field
                for field, normalized in field_map.items()
                if "date/time" in normalized or ("date" in normalized and "time" in normalized)
            ),
            None,
        )
        value_field = next(
            (
                field
                for field, normalized in field_map.items()
                if normalized == "value" or normalized.endswith("value")
            ),
            None,
        )
        unit_field = next(
            (
                field
                for field, normalized in field_map.items()
                if normalized == "unit" or normalized.endswith("unit")
            ),
            None,
        )
        if datetime_field is None or value_field is None:
            raise AppError(
                "CSV_HEADER_INVALID",
                "CSV 文件缺少 Date/Time 或 Value 列",
                detail=str(path),
            )

        records: list[dict[str, Any]] = []
        point_index = 0
        first_unit_value: str | None = None
        for row_index, row in enumerate(reader, start=2):
            dt_raw = _safe_str(row.get(datetime_field))
            value_raw = row.get(value_field)
            if not dt_raw:
                continue
            point_index += 1
            if first_unit_value is None and unit_field is not None:
                unit_text = _safe_str(row.get(unit_field))
                if unit_text:
                    first_unit_value = unit_text
            try:
                dt_value = datetime.strptime(dt_raw, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                match = re.search(
                    r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})[:：](\d{1,2})[:：](\d{1,2})",
                    dt_raw,
                )
                if not match:
                    raise AppError(
                        "CSV_DATETIME_INVALID",
                        "CSV 时间格式无法解析",
                        detail=f"{path.name} row={row_index} value={dt_raw}",
                    )
                dt_value = datetime(*map(int, match.groups()))

            records.append(
                {
                    "point_index": point_index,
                    "dt": dt_value,
                    "value": value_raw,
                }
            )

    if not records:
        raise AppError("CSV_NO_DATA", "CSV 文件没有可用数据", detail=str(path))

    records, warnings = _normalize_duplicate_second_records(records, "CSV", path)
    mapping = {record["dt"]: record["value"] for record in records}
    return mapping, warnings, first_unit_value


def parse_voltage_csv(path: Path) -> tuple[dict[datetime, float], list[str]]:
    raw_mapping, warnings, _ = parse_time_value_csv(path)
    mapping: dict[datetime, float] = {}
    for dt_value, raw_value in raw_mapping.items():
        numeric_value = _to_float(raw_value)
        if numeric_value is None:
            raise AppError(
                "CSV_VALUE_INVALID",
                "CSV Value 列存在无法解析的数据",
                detail=f"{path.name} value={raw_value}",
            )
        mapping[dt_value] = numeric_value
    return mapping, warnings
