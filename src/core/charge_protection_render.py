from __future__ import annotations

import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ChargeProtectionDataset, ChargeProtectionMetrics


DEFAULT_CHART_WIDTH = 16
DEFAULT_CHART_HEIGHT = 10
PLOT_AREA_X = 0.0
PLOT_AREA_Y = 0.0
PLOT_AREA_W = 0.9
PLOT_AREA_H = 0.7
CHART_TITLE_SIZE = 1600
MISSING_VALUE_TEXT = "未检测到符合要求的数据，请人工查看"


@dataclass(slots=True)
class _ColumnDef:
    header: str
    values: list[Any]
    number_format: str | None = None


def _format_temperature(value: float | None) -> str:
    if value is None:
        return MISSING_VALUE_TEXT
    return f"{value:.3f} °C"


def _write_summary_table(
    ws: Any,
    *,
    start_col: int,
    title: str,
    rows: list[tuple[str, str]],
    start_row: int = 1,
) -> int:
    title_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 1,
    )
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = title_font
    title_cell.alignment = center

    header_left = ws.cell(row=start_row + 1, column=start_col, value="测试项")
    header_right = ws.cell(row=start_row + 1, column=start_col + 1, value="数据")
    for cell in (header_left, header_right):
        cell.font = title_font
        cell.alignment = center
        cell.fill = header_fill

    for offset, (name, value) in enumerate(rows, start=2):
        ws.cell(row=start_row + offset, column=start_col, value=name)
        value_cell = ws.cell(row=start_row + offset, column=start_col + 1, value=value)
        value_cell.alignment = center

    ws.column_dimensions[get_column_letter(start_col)].width = 20
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 28
    return start_row + 1 + len(rows)


def _apply_plot_area_layout(chart: LineChart) -> None:
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="outer",
            x=PLOT_AREA_X,
            y=PLOT_AREA_Y,
            w=PLOT_AREA_W,
            h=PLOT_AREA_H,
            xMode="factor",
            yMode="factor",
            wMode="factor",
            hMode="factor",
        )
    )


def _add_protection_chart(
    ws: Any,
    *,
    anchor_cell: str,
    title: str,
    first_data_row: int,
    last_data_row: int,
    time_col: int,
    current_col: int,
    voltage_col: int,
    env_col: int,
    cell_col: int,
) -> None:
    primary_x_axis_id = 500
    primary_y_axis_id = 100
    secondary_y_axis_id = 200

    chart_left = LineChart()
    chart_left.title = title
    chart_left.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=CHART_TITLE_SIZE, b=True)
    chart_left.x_axis.axId = primary_x_axis_id
    chart_left.y_axis.axId = primary_y_axis_id
    chart_left.y_axis.title = "电流"
    chart_left.y_axis.axPos = "l"
    chart_left.y_axis.delete = False
    chart_left.y_axis.tickLblPos = "nextTo"
    chart_left.y_axis.number_format = "0.000"
    chart_left.y_axis.crosses = "min"
    chart_left.y_axis.crossAx = primary_x_axis_id
    chart_left.x_axis.title = "时间"
    chart_left.x_axis.axPos = "b"
    chart_left.x_axis.delete = False
    chart_left.x_axis.crossAx = primary_y_axis_id
    chart_left.x_axis.number_format = "hh:mm:ss"
    chart_left.x_axis.tickLblPos = "low"
    chart_left.legend.position = "b"
    chart_left.width = DEFAULT_CHART_WIDTH
    chart_left.height = DEFAULT_CHART_HEIGHT
    _apply_plot_area_layout(chart_left)

    current_data = Reference(ws, min_col=current_col, min_row=1, max_col=current_col, max_row=last_data_row)
    categories = Reference(ws, min_col=time_col, min_row=first_data_row, max_col=time_col, max_row=last_data_row)
    chart_left.add_data(current_data, titles_from_data=True)
    chart_left.set_categories(categories)

    chart_right = LineChart()
    chart_right.x_axis.axId = primary_x_axis_id
    chart_right.y_axis.axId = secondary_y_axis_id
    chart_right.y_axis.title = "温度"
    chart_right.y_axis.axPos = "r"
    chart_right.y_axis.delete = False
    chart_right.y_axis.tickLblPos = "nextTo"
    chart_right.y_axis.number_format = "0.000"
    chart_right.y_axis.crosses = "max"
    chart_right.y_axis.crossAx = primary_x_axis_id
    chart_right.y_axis.majorGridlines = None
    chart_right.width = DEFAULT_CHART_WIDTH
    chart_right.height = DEFAULT_CHART_HEIGHT
    for series_col in (env_col, cell_col, voltage_col):
        chart_right.add_data(
            Reference(ws, min_col=series_col, min_row=1, max_col=series_col, max_row=last_data_row),
            titles_from_data=True,
        )
    chart_left += chart_right
    ws.add_chart(chart_left, anchor_cell)


def _center_cells_with_data(ws: Any) -> None:
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.strip() == "":
                continue
            cell.alignment = center


def _build_summary_content(metrics: ChargeProtectionMetrics) -> tuple[str, list[tuple[str, str]]]:
    if metrics.mode == "high_low":
        return (
            "高低温充电保护测试",
            [
                ("高温充电保护温度", _format_temperature(metrics.high_protect_temp_c)),
                ("高温充电复充温度", _format_temperature(metrics.high_resume_temp_c)),
                ("低温充电保护温度", _format_temperature(metrics.low_protect_temp_c)),
                ("低温充电复充温度", _format_temperature(metrics.low_resume_temp_c)),
            ],
        )
    if metrics.mode == "high":
        return (
            "高温充电保护测试",
            [
                ("高温充电保护温度", _format_temperature(metrics.high_protect_temp_c)),
                ("高温充电复充温度", _format_temperature(metrics.high_resume_temp_c)),
            ],
        )
    return (
        "低温充电保护测试",
        [
            ("低温充电保护温度", _format_temperature(metrics.low_protect_temp_c)),
            ("低温充电复充温度", _format_temperature(metrics.low_resume_temp_c)),
        ],
    )


def _save_workbook(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_dir = output_path.parent / ".temp"
    temp_dir.mkdir(exist_ok=True)

    original_tempdir = tempfile.tempdir
    try:
        tempfile.tempdir = str(temp_dir)
        workbook.save(output_path)
    finally:
        tempfile.tempdir = original_tempdir
        try:
            for temp_file in temp_dir.glob("openpyxl.*"):
                temp_file.unlink(missing_ok=True)
            if not any(temp_dir.iterdir()):
                temp_dir.rmdir()
        except Exception:
            pass


def render_charge_protection_workbook(
    dataset: ChargeProtectionDataset,
    metrics: ChargeProtectionMetrics,
    output_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统计结果"

    columns: list[_ColumnDef] = []
    if any(value is not None for value in dataset.index_values):
        columns.append(_ColumnDef("索引", dataset.index_values))
    columns.append(_ColumnDef("日期", dataset.date_strings))
    columns.append(_ColumnDef("时间 (s)", dataset.datetimes, number_format="hh:mm:ss"))
    columns.append(_ColumnDef("电流 (mA)", dataset.currents_ma, number_format="0.000"))
    columns.append(_ColumnDef("电压 (V)", dataset.voltages_v, number_format="0.000"))
    columns.append(_ColumnDef("环境温度 (°C)", dataset.env_temps_c, number_format="0.000"))
    columns.append(_ColumnDef("电芯温度 (°C)", dataset.cell_temps_c, number_format="0.000"))

    header_font = Font(bold=True)
    for col_idx, column in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_idx, value=column.header).font = header_font
        if column.header in {"日期", "时间 (s)"}:
            width = 14
        elif "温度" in column.header:
            width = 16
        else:
            width = 14
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset in range(dataset.row_count()):
        row_idx = row_offset + 2
        for col_idx, column in enumerate(columns, start=1):
            value = column.values[row_offset] if row_offset < len(column.values) else None
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if column.number_format and value is not None:
                cell.number_format = column.number_format

    summary_start_col = len(columns) + 2
    summary_title, summary_rows = _build_summary_content(metrics)
    summary_end_row = _write_summary_table(sheet, start_col=summary_start_col, title=summary_title, rows=summary_rows)

    header_to_col = {column.header: idx for idx, column in enumerate(columns, start=1)}
    chart_anchor = f"{get_column_letter(summary_start_col)}{summary_end_row + 2}"
    _add_protection_chart(
        sheet,
        anchor_cell=chart_anchor,
        title=dataset.stem,
        first_data_row=2,
        last_data_row=dataset.row_count() + 1,
        time_col=header_to_col["时间 (s)"],
        current_col=header_to_col["电流 (mA)"],
        voltage_col=header_to_col["电压 (V)"],
        env_col=header_to_col["环境温度 (°C)"],
        cell_col=header_to_col["电芯温度 (°C)"],
    )

    _center_cells_with_data(sheet)
    _save_workbook(workbook, output_path)
