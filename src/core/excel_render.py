from __future__ import annotations

from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ChargeDataset, ChargeMetrics, TemperatureMetrics


DEFAULT_CHART_WIDTH = 16
DEFAULT_CHART_HEIGHT = 10
TEMP_CHART_ROW_GAP = 22

# Shrink only the plot area (line region) while keeping the chart frame size unchanged.
PLOT_AREA_X = 0.0
PLOT_AREA_Y = 0.0
PLOT_AREA_W = 0.9
PLOT_AREA_H = 0.7
CHART_TITLE_SIZE = 1600


@dataclass(slots=True)
class _ColumnDef:
    header: str
    values: list[Any]
    number_format: str | None = None


def _format_float(value: float | None, unit: str = "") -> str:
    if value is None:
        return "未检测到符合要求的数据，请人工查看"
    if unit:
        return f"{value:.3f} {unit}"
    return f"{value:.3f}"


def _format_duration(value: timedelta | None) -> str:
    if value is None:
        return "未检测到符合要求的数据，请人工查看"
    total_seconds = int(value.total_seconds())
    if total_seconds < 0:
        total_seconds = abs(total_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def _curve_chart_title(stem: str) -> str:
    if "充电曲线" in stem:
        return stem
    if "充电温升" in stem:
        return stem.replace("温升", "曲线")
    return "充电曲线测试"


def _temp_chart_title(stem: str) -> str:
    if "充电温升" in stem:
        return stem
    if "充电曲线" in stem:
        return stem.replace("曲线", "温升")
    return "充电温升测试"


def _write_summary_table(
    ws: Any,
    start_col: int,
    title: str,
    rows: list[tuple[str, str]],
    start_row: int = 1,
) -> int:
    title_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 1,
    )
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = title_font
    title_cell.alignment = center

    ws.cell(row=start_row + 1, column=start_col, value="测试项").font = title_font
    ws.cell(row=start_row + 1, column=start_col + 1, value="数据").font = title_font
    ws.cell(row=start_row + 1, column=start_col).alignment = center
    ws.cell(row=start_row + 1, column=start_col + 1).alignment = center
    ws.cell(row=start_row + 1, column=start_col).fill = header_fill
    ws.cell(row=start_row + 1, column=start_col + 1).fill = header_fill

    for offset, (name, value) in enumerate(rows, start=2):
        ws.cell(row=start_row + offset, column=start_col, value=name)
        ws.cell(row=start_row + offset, column=start_col + 1, value=value)
        ws.cell(row=start_row + offset, column=start_col + 1).alignment = center
    ws.column_dimensions[get_column_letter(start_col)].width = 16
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 28
    return start_row + 1 + len(rows)


def _apply_plot_area_layout(chart: LineChart) -> None:
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="outer",
            x=PLOT_AREA_X,
            y=PLOT_AREA_Y,
            w=PLOT_AREA_W,
            h=PLOT_AREA_H,
            xMode="factor",
            yMode="factor",
            wMode="factor",
            hMode="factor",
        )
    )


def _add_curve_chart(
    ws: Any,
    anchor_cell: str,
    title: str,
    first_data_row: int,
    last_data_row: int,
    time_col: int | None,
    current_col: int,
    voltage_col: int | None,
) -> None:
    primary_x_axis_id = 500
    primary_y_axis_id = 100
    secondary_y_axis_id = 200

    chart_left = LineChart()
    chart_left.title = title
    chart_left.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=CHART_TITLE_SIZE, b=True)
    chart_left.x_axis.axId = primary_x_axis_id
    chart_left.y_axis.title = "电流"
    chart_left.y_axis.axId = primary_y_axis_id
    chart_left.y_axis.axPos = "l"
    chart_left.y_axis.delete = False
    chart_left.y_axis.tickLblPos = "nextTo"
    chart_left.y_axis.number_format = "0.000"
    chart_left.y_axis.crosses = "min"
    chart_left.y_axis.crossAx = primary_x_axis_id
    chart_left.x_axis.title = "时间"
    chart_left.x_axis.axPos = "b"
    chart_left.x_axis.delete = False
    chart_left.x_axis.crossAx = primary_y_axis_id
    # Keep X axis as native Excel time to avoid Office/WPS rendering differences.
    chart_left.x_axis.number_format = "hh:mm:ss"
    chart_left.x_axis.tickLblPos = "low"
    chart_left.legend.position = "b"
    chart_left.width = DEFAULT_CHART_WIDTH
    chart_left.height = DEFAULT_CHART_HEIGHT
    _apply_plot_area_layout(chart_left)

    current_data = Reference(
        ws,
        min_col=current_col,
        min_row=1,
        max_col=current_col,
        max_row=last_data_row,
    )
    chart_left.add_data(current_data, titles_from_data=True)
    if time_col is not None:
        categories = Reference(
            ws,
            min_col=time_col,
            min_row=first_data_row,
            max_col=time_col,
            max_row=last_data_row,
        )
        chart_left.set_categories(categories)

    if voltage_col is not None:
        chart_right = LineChart()
        chart_right.x_axis.axId = primary_x_axis_id
        chart_right.y_axis.axId = secondary_y_axis_id
        chart_right.y_axis.title = "电压"
        chart_right.y_axis.axPos = "r"
        chart_right.y_axis.delete = False
        chart_right.y_axis.tickLblPos = "nextTo"
        chart_right.y_axis.number_format = "0.000"
        chart_right.y_axis.crosses = "max"
        chart_right.y_axis.crossAx = primary_x_axis_id
        chart_right.y_axis.majorGridlines = None
        chart_right.width = DEFAULT_CHART_WIDTH
        chart_right.height = DEFAULT_CHART_HEIGHT
        voltage_data = Reference(
            ws,
            min_col=voltage_col,
            min_row=1,
            max_col=voltage_col,
            max_row=last_data_row,
        )
        chart_right.add_data(voltage_data, titles_from_data=True)
        chart_left += chart_right
    ws.add_chart(chart_left, anchor_cell)


def _add_temp_chart(
    ws: Any,
    anchor_cell: str,
    title: str,
    first_data_row: int,
    last_data_row: int,
    time_col: int | None,
    pen_col: int,
    env_col: int,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=CHART_TITLE_SIZE, b=True)
    chart.x_axis.title = "时间"
    chart.x_axis.axPos = "b"
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = "温度"
    chart.y_axis.axPos = "l"
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.number_format = "0.000"
    chart.y_axis.crosses = "min"
    chart.legend.position = "b"
    chart.width = DEFAULT_CHART_WIDTH
    chart.height = DEFAULT_CHART_HEIGHT
    _apply_plot_area_layout(chart)
    pen_data = Reference(ws, min_col=pen_col, min_row=1, max_col=pen_col, max_row=last_data_row)
    env_data = Reference(ws, min_col=env_col, min_row=1, max_col=env_col, max_row=last_data_row)
    chart.add_data(pen_data, titles_from_data=True)
    chart.add_data(env_data, titles_from_data=True)
    if time_col is not None:
        categories = Reference(
            ws,
            min_col=time_col,
            min_row=first_data_row,
            max_col=time_col,
            max_row=last_data_row,
        )
        chart.set_categories(categories)
    ws.add_chart(chart, anchor_cell)


def _center_cells_with_data(ws: Any) -> None:
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and value.strip() == "":
                continue
            cell.alignment = center


def render_charge_workbook(
    dataset: ChargeDataset,
    charge_metrics: ChargeMetrics,
    temperature_metrics: TemperatureMetrics | None,
    output_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统计结果"

    columns: list[_ColumnDef] = []
    if any(value is not None for value in dataset.index_values):
        columns.append(_ColumnDef("索引", dataset.index_values))
    columns.append(_ColumnDef("日期", dataset.date_strings))
    columns.append(_ColumnDef("时间 (s)", dataset.datetimes, number_format="hh:mm:ss"))
    columns.append(_ColumnDef("电流 (mA)", dataset.currents_ma, number_format="0.000"))
    if any(value is not None for value in dataset.voltages_v):
        columns.append(_ColumnDef("电压（V）", dataset.voltages_v, number_format="0.000"))
    if dataset.has_temperature_data:
        columns.append(_ColumnDef("笔壳温度 (°C)", dataset.pen_temps_c, number_format="0.000"))
        columns.append(_ColumnDef("环境温度 (°C)", dataset.env_temps_c, number_format="0.000"))
    for header in dataset.extra_headers_order:
        values = dataset.extras.get(header, [])
        columns.append(_ColumnDef(header, values))

    header_font = Font(bold=True)
    for col_idx, column in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_idx, value=column.header).font = header_font
        if column.header in ("日期", "时间 (s)"):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 14
        elif "温度" in column.header:
            sheet.column_dimensions[get_column_letter(col_idx)].width = 16
        else:
            sheet.column_dimensions[get_column_letter(col_idx)].width = 14

    for row_offset in range(dataset.row_count()):
        row_idx = row_offset + 2
        for col_idx, column in enumerate(columns, start=1):
            value = column.values[row_offset] if row_offset < len(column.values) else None
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if column.number_format and value is not None:
                cell.number_format = column.number_format

    curve_table_col = len(columns) + 2
    curve_rows = [
        ("预充电流", _format_float(charge_metrics.precharge_current_ma, "mA")),
        ("恒流电流", _format_float(charge_metrics.const_current_ma, "mA")),
        ("截充电流", _format_float(charge_metrics.cutoff_current_ma, "mA")),
        ("满充电压", _format_float(charge_metrics.full_voltage_v, "V")),
        ("充电时长", _format_duration(charge_metrics.duration)),
    ]
    curve_table_end_row = _write_summary_table(sheet, curve_table_col, "充电曲线测试", curve_rows)

    temp_table_end_row = 0
    if dataset.has_temperature_data and temperature_metrics is not None:
        temp_table_col = curve_table_col + 3
        temp_rows = [
            ("笔壳最高温度", _format_float(temperature_metrics.max_pen_temp_c, "°C")),
            ("对应的环境温度", _format_float(temperature_metrics.env_temp_at_max_pen_c, "°C")),
            ("热点处温升", _format_float(temperature_metrics.hotspot_rise_c, "°C")),
        ]
        temp_table_end_row = _write_summary_table(
            sheet,
            temp_table_col,
            "充电温升测试",
            temp_rows,
        )

    data_start_row = 2
    data_end_row = dataset.row_count() + 1
    header_to_col = {column.header: idx for idx, column in enumerate(columns, start=1)}
    time_col = header_to_col.get("时间 (s)")
    current_col = header_to_col["电流 (mA)"]
    voltage_col = header_to_col.get("电压（V）")
    chart_start_row = max(curve_table_end_row, temp_table_end_row or 0) + 2
    chart_anchor = f"{get_column_letter(curve_table_col)}{chart_start_row}"
    _add_curve_chart(
        sheet,
        chart_anchor,
        _curve_chart_title(dataset.stem),
        data_start_row,
        data_end_row,
        time_col,
        current_col,
        voltage_col,
    )

    if dataset.has_temperature_data and temperature_metrics is not None:
        pen_col = header_to_col["笔壳温度 (°C)"]
        env_col = header_to_col["环境温度 (°C)"]
        temp_anchor_row = chart_start_row + TEMP_CHART_ROW_GAP
        temp_anchor = f"{get_column_letter(curve_table_col)}{temp_anchor_row}"
        _add_temp_chart(
            sheet,
            temp_anchor,
            _temp_chart_title(dataset.stem),
            data_start_row,
            data_end_row,
            time_col,
            pen_col,
            env_col,
        )

    _center_cells_with_data(sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
